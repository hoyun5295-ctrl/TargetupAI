# -*- coding: utf-8 -*-
"""
30,000건 가상 CRM 고객 엑셀 파일 생성 (D210+ Phase 3 테스트 영역)
- 출력: C:\\Users\\ceo\\projects\\targetup\\test_customers_30000.xlsx
- 전화번호: 010-0001-0001 ~ 010-0003-0000 (안전 시퀀스 — 실 발송 차단)
- 현재 날짜 2026-05-23 기준 분포 매트릭스
"""

import random
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

random.seed(42)

# ── 매트릭스 영역 ──────────────────────────────────────
SURNAMES = ['김','이','박','최','정','강','조','윤','장','임','한','오','서','신','권','황','안','송','류','전','홍','고','문','양','손','배','백','허','유','남']
MALE_NAMES = ['민준','서준','도윤','예준','시우','하준','지호','지후','준서','준우','현우','도현','지훈','건우','우진','선우','서진','민재','현준','연우','정우','승우','승현','시윤','진우','지환','수호','지원','승준','시원']
FEMALE_NAMES = ['서연','서윤','지우','서현','민서','하은','하윤','윤서','지유','지민','채원','수아','지아','지윤','은서','다은','예은','수빈','소율','예린','예원','시아','지안','채은','유나','윤아','시은','다인','서아','하린']
EMAIL_DOMAINS = ['gmail.com','naver.com','kakao.com','daum.net','hanmail.net']
STREETS = ['중앙로','강남대로','테헤란로','을지로','종로','가락로','학동로','반포대로','광교로','월드컵로']
CATEGORIES = ['패션','뷰티','식품','생활용품','전자제품','스포츠','도서','유아용품','반려동물','홈데코']
OCCUPATIONS = ['직장인','자영업','전업주부','학생','프리랜서','공무원','전문직','기타']
BRANDS = ['샤넬','구찌','루이비통','프라다','에르메스','버버리','디올','발렌시아가','펜디','지방시','없음']

STORES = {
    'STORE_001': '강남점', 'STORE_002': '홍대점', 'STORE_003': '명동점',
    'STORE_004': '잠실점', 'STORE_005': '신촌점', 'STORE_006': '건대점',
    'STORE_007': '여의도점', 'STORE_008': '판교점', 'STORE_009': '분당점',
    'STORE_010': '일산점', 'STORE_011': '부산서면점', 'STORE_012': '해운대점',
    'STORE_013': '대구동성로점', 'STORE_014': '인천논현점', 'STORE_015': '광주충장로점',
    'STORE_016': '대전둔산점', 'STORE_017': '울산삼산점', 'STORE_018': '청주성안길점',
    'STORE_019': '전주신시가지점', 'STORE_020': '제주노형점'
}

REFERENCE_DATE = date(2026, 5, 23)

# ── 가중치 매트릭스 ────────────────────────────────────
GRADE_WEIGHTS = [('VIP', 5), ('Gold', 10), ('Silver', 20), ('Bronze', 25), ('일반', 30), ('신규', 10)]
REGION_WEIGHTS = [
    ('서울', 25), ('경기', 22), ('부산', 8), ('인천', 6), ('대구', 5), ('대전', 4),
    ('광주', 4), ('울산', 3), ('세종', 1), ('강원', 5), ('충북', 4), ('충남', 4),
    ('전북', 4), ('전남', 2), ('경북', 4), ('경남', 1), ('제주', 2)
]
REGISTRATION_WEIGHTS = [('오프라인', 50), ('온라인', 35), ('이벤트', 10), ('제휴', 5)]

GRADE_PURCHASE_AMOUNT = {
    'VIP': (5000000, 20000000), 'Gold': (2000000, 5000000),
    'Silver': (500000, 2000000), 'Bronze': (100000, 500000),
    '일반': (10000, 100000), '신규': (0, 0)
}
GRADE_PURCHASE_COUNT = {
    'VIP': (50, 200), 'Gold': (20, 50), 'Silver': (10, 25),
    'Bronze': (3, 10), '일반': (1, 4), '신규': (0, 0)
}
GRADE_LTV = {
    'VIP': (80, 100), 'Gold': (60, 80), 'Silver': (40, 60),
    'Bronze': (20, 40), '일반': (5, 20), '신규': (0, 0)
}
GRADE_POINTS = {
    'VIP': (50000, 100000), 'Gold': (20000, 50000),
    'Silver': (5000, 20000), 'Bronze': (1000, 5000),
    '일반': (0, 1000), '신규': (0, 0)
}


def weighted_choice(weights_list):
    items = [w[0] for w in weights_list]
    weights = [w[1] for w in weights_list]
    return random.choices(items, weights=weights, k=1)[0]


def random_birth_year():
    r = random.random()
    if r < 0.25: return random.randint(1996, 2005)  # 20대
    elif r < 0.55: return random.randint(1986, 1995)  # 30대
    elif r < 0.75: return random.randint(1976, 1985)  # 40대
    elif r < 0.90: return random.randint(1966, 1975)  # 50대
    else: return random.randint(1950, 1965)  # 60대+


def random_recent_purchase_date():
    r = random.random()
    if r < 0.05: days_ago = random.randint(0, 7)
    elif r < 0.20: days_ago = random.randint(8, 30)
    elif r < 0.45: days_ago = random.randint(31, 90)
    elif r < 0.70: days_ago = random.randint(91, 180)
    elif r < 0.90: days_ago = random.randint(181, 365)
    else: days_ago = random.randint(366, 730)
    return REFERENCE_DATE - timedelta(days=days_ago)


# ── 워크북 생성 ────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = '고객데이터'

headers = [
    '휴대폰번호', '이름', '성별', '생년월일', '나이',
    '이메일', '주소', '지역', '등급',
    '매장코드', '매장명', '가입경로',
    '포인트', '최근구매일', '최근구매금액',
    '누적구매금액', '구매횟수', '평균객단가', 'LTV점수',
    '결혼기념일', 'SMS수신동의', '활성여부',
    '선호카테고리', '직업', '자녀수', '선호브랜드'
]
ws.append(headers)

# 헤더 스타일 영역
header_font = Font(bold=True, color='FFFFFF', name='맑은 고딕')
header_fill = PatternFill('solid', start_color='4472C4')
header_alignment = Alignment(horizontal='center', vertical='center')
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment

# ── 30,000 row 생성 ───────────────────────────────────
TOTAL_ROWS = 30000
for n in range(1, TOTAL_ROWS + 1):
    first_part = (n - 1) // 10000 + 1
    second_part = (n - 1) % 10000 + 1
    phone = f'010-{first_part:04d}-{second_part:04d}'

    is_male = random.random() < 0.5
    surname = random.choice(SURNAMES)
    firstname = random.choice(MALE_NAMES if is_male else FEMALE_NAMES)
    name = surname + firstname
    gender = '남' if is_male else '여'

    birth_year = random_birth_year()
    birth_month = random.randint(1, 12)
    birth_day = random.randint(1, 28)
    birth_date_str = f'{birth_year:04d}-{birth_month:02d}-{birth_day:02d}'
    age = 2026 - birth_year

    email_domain = random.choice(EMAIL_DOMAINS)
    email = f'test{n}@{email_domain}'

    region = weighted_choice(REGION_WEIGHTS)
    street = random.choice(STREETS)
    address = f'{region} {street} {random.randint(1, 999)}-{random.randint(1, 99)}'

    grade = weighted_choice(GRADE_WEIGHTS)

    store_code = random.choice(list(STORES.keys()))
    store_name = STORES[store_code]

    registration_type = weighted_choice(REGISTRATION_WEIGHTS)

    pts_min, pts_max = GRADE_POINTS[grade]
    points = random.randint(pts_min, pts_max) if pts_max > 0 else 0

    recent_purchase_date = random_recent_purchase_date()
    recent_purchase_date_str = recent_purchase_date.strftime('%Y-%m-%d')

    if grade == '신규':
        recent_purchase_amount = 0
    else:
        recent_purchase_amount = random.randint(10000, 500000)

    amt_min, amt_max = GRADE_PURCHASE_AMOUNT[grade]
    total_purchase_amount = random.randint(amt_min, amt_max) if amt_max > 0 else 0

    cnt_min, cnt_max = GRADE_PURCHASE_COUNT[grade]
    purchase_count = random.randint(cnt_min, cnt_max) if cnt_max > 0 else 0

    avg_order_value = total_purchase_amount // purchase_count if purchase_count > 0 else 0

    ltv_min, ltv_max = GRADE_LTV[grade]
    ltv_score = random.randint(ltv_min, ltv_max) if ltv_max > 0 else 0

    is_married = random.random() < 0.40
    if is_married:
        years_married = random.randint(3, 18)
        wedding_date = REFERENCE_DATE - timedelta(days=365 * years_married)
        wedding_date_str = wedding_date.strftime('%Y-%m-%d')
    else:
        wedding_date_str = ''

    sms_opt_in = 'Y' if random.random() < 0.95 else 'N'
    is_active = 'Y' if random.random() < 0.90 else 'N'

    preferred_category = random.choice(CATEGORIES)
    occupation = random.choice(OCCUPATIONS)
    children_count = random.randint(0, 3)
    favorite_brand = random.choice(BRANDS)

    row = [
        phone, name, gender, birth_date_str, age,
        email, address, region, grade,
        store_code, store_name, registration_type,
        points, recent_purchase_date_str, recent_purchase_amount,
        total_purchase_amount, purchase_count, avg_order_value, ltv_score,
        wedding_date_str, sms_opt_in, is_active,
        preferred_category, occupation, children_count, favorite_brand
    ]
    ws.append(row)

# ── 컬럼 너비 ──────────────────────────────────────────
column_widths = {
    'A': 16, 'B': 10, 'C': 6, 'D': 12, 'E': 6,
    'F': 28, 'G': 28, 'H': 8, 'I': 8,
    'J': 12, 'K': 16, 'L': 10,
    'M': 10, 'N': 12, 'O': 12,
    'P': 14, 'Q': 8, 'R': 12, 'S': 8,
    'T': 12, 'U': 12, 'V': 8,
    'W': 10, 'X': 8, 'Y': 8, 'Z': 10
}
for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

ws.freeze_panes = 'A2'

# ── 데이터 영역 폰트 통일 ──────────────────────────────
data_font = Font(name='맑은 고딕', size=10)
for row in ws.iter_rows(min_row=2, max_row=TOTAL_ROWS + 1):
    for cell in row:
        cell.font = data_font

# ── 저장 ───────────────────────────────────────────────
output_path = r'C:\Users\ceo\projects\targetup\test_customers_30000.xlsx'
wb.save(output_path)

print(f'엑셀 파일 작성 완료: {output_path}')
print(f'행 수: {TOTAL_ROWS:,} + 1 (헤더)')
print(f'컬럼 수: {len(headers)}개')
print(f'전화번호 영역: 010-0001-0001 ~ 010-0003-0000')

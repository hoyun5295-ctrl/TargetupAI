# -*- coding: utf-8 -*-
# imweb 앱스토어 앱 소개 문구 양식 채우기 — A열 라벨 / B열 답변 구조
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

SRC = r"C:\Users\ceo\projects\targetup\docs\imweb-appstore\build\imweb_form.xlsx"
OUT = r"C:\Users\ceo\projects\targetup\docs\imweb-appstore\deliverables\imweb_form_hanjulAI_filled.xlsx"

INTRO = """한줄로AI는 아임웹 사이트의 회원·주문 데이터를 자동으로 연동해, AI가 타겟을 찾고 문자·알림톡·이메일·모바일 DM·인앱 메시지까지 만들어 보내는 마케팅 자동화 서비스입니다.

[이런 점이 다릅니다]
- 연동 1분: 앱 추가와 권한 동의만 하면 회원·주문·장바구니 데이터가 실시간으로 들어옵니다. 개발 작업이 필요 없습니다.
- AI가 콘텐츠를 만듭니다: 상품 사진 한 장이면 배경·연출·타이포까지 완성된 포스터가 나옵니다(이미지 스튜디오, 7개 카테고리 58종 템플릿).
- 타겟을 알아서 찾습니다: 구매 이력·등급·행동 데이터 기반으로 보낼 사람을 자동 추출합니다.
- 시점도 알아서: 가입·구매·이탈 같은 고객 행동 시점에 맞춰 자동 발송되는 AI 여정을 만들 수 있습니다.
- 성과가 보입니다: 발송 결과와 매출 기여를 대시보드에서 확인합니다.

[주요 기능]
1. 이미지 스튜디오 — 템플릿 고르고 상품·문구만 넣으면 완성 포스터. 만든 소재는 문자(MMS)·DM·이메일·인앱에 바로 사용.
2. 모바일 DM·이메일 편집기 — 연동 몰에서 상품을 불러오면 사진·가격·할인율이 자동으로 채워집니다.
3. 인앱 메시지 — 사이트 방문자에게 띄우는 메시지를 실시간 미리보기로 편집.
4. AI 여정 자동화 — 시점·타겟·문구를 AI가 설계하는 자동 캠페인.
5. 발송 채널 — 문자(SMS·LMS·MMS) · 카카오 알림톡 · 이메일 · 모바일 DM · 인앱 메시지.

[요금 안내]
앱 설치와 아임웹 연동은 무료입니다. 한줄로 서비스는 30일 무료체험 후 월 15만 원부터 시작하는 구독 요금제로 이용할 수 있으며, 문자 발송 요금은 별도입니다. 자세한 내용은 hanjul.ai에서 확인하세요."""

FILLS = {
    "B1": "한줄로AI",
    "B2": "아임웹 회원·주문을 자동 연동해 문자·카카오·DM·인앱까지 AI 타겟 마케팅",
    "B3": "https://hanjul.ai",
    "B4": INTRO,
    "B8": "설치하면 어떤 데이터가 연동되나요?",
    "B9": "회원 정보(수신 동의 포함), 주문, 장바구니, 배송 상태가 연동됩니다. 연동 항목은 권한 동의 화면에서 확인할 수 있으며, 연동 해제 시 즉시 수집이 중단됩니다.",
    "B11": "개발자가 없어도 쓸 수 있나요?",
    "B12": "네. 아임웹 앱스토어에서 앱 추가와 동의만 하면 연동이 끝나고, 이후 모든 기능은 관리 화면에서 클릭으로 사용합니다.",
    "B14": "문의는 어디로 하나요?",
    "B15": "mobile@invitocorp.com 으로 문의해 주세요. 홈페이지(hanjul.ai)에서도 안내를 확인할 수 있습니다.",
}

wb = openpyxl.load_workbook(SRC)
ws = wb.active
label_font = ws["A1"].font
base = Font(name=label_font.name or "맑은 고딕", size=label_font.size or 11)
for coord, val in FILLS.items():
    c = ws[coord]
    c.value = val
    c.font = base
    c.alignment = Alignment(wrap_text=True, vertical="top")
ws.column_dimensions["B"].width = 95
ws.row_dimensions[4].height = 420
wb.save(OUT)
print("saved:", OUT)

# 검증 재로드 — 채운 값 회독
wb2 = openpyxl.load_workbook(OUT)
ws2 = wb2.active
for coord in FILLS:
    v = ws2[coord].value
    print(coord, "OK" if v else "EMPTY", str(v)[:40].replace("\n", " "))
